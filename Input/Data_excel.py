import openpyxl


def row_count():
    path = "..\\Input\\Data.xlsx"
    wb = openpyxl.load_workbook(path)

    # sheet = wb.active
    sheet = wb.get_sheet_by_name("Sheet1")

    #  how many no. of rows and coloumn
    rows = sheet.max_row
    column = sheet.max_column

    print(rows)
    print(column)

    for r in range(1, rows + 1):
        for c in range(1, column + 1):
            print(sheet.cell(row=r, column=c).value, end="        ")

        # To go next line for outer loop
        print()


row_count()
